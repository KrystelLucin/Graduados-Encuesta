import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart,Reference

def existe(identificacion):
    wb = openpyxl.load_workbook('data/FIEC Graduados_Encuesta_2018_2020.xlsx')
    alumnosFIEC = wb.get_sheet_by_name('FIEC')
    for row in alumnosFIEC.iter_rows():
        if row[5].value == identificacion:
            carrera = row[2].value
            if carrera == 'Computación':
                set_respondido_computacion(identificacion)
            else:
                if carrera == 'Electrónica y Automatización' or carrera == 'Telemática' :
                    set_respondido(identificacion, 'Electronica')
                else:
                    if carrera == 'Telemática':
                        set_respondido(identificacion, 'Telematica')
                    else:
                        if carrera is not None:
                            set_respondido(identificacion, carrera)
            return True
    return False

def set_respondido_computacion(identificacion):
    wb = openpyxl.load_workbook('data/FIEC_Computacion_2018_2020.xlsx')
    alumnos = wb['Hoja1']
    respondido = alumnos['P']

    encabezado = True
    for celda in respondido:
        if encabezado == False:
            if celda.value != 'SI' and alumnos[f"H{celda.row}"].value == identificacion:
                alumnos[f"P{celda.row}"] = 'SI'
                celda.fill = PatternFill("solid", fgColor="92D050")
                wb.save("data/FIEC_Computacion_2018_2020.xlsx")
                return True
        else:
            encabezado = False


def set_respondido(identificacion, carrera):
    wb = openpyxl.load_workbook('data/FIEC_%s_2018_2020.xlsx' % carrera)
    alumnos = wb['Hoja1']
    respondido = alumnos['P']

    encabezado = True
    for celda in respondido:
        if encabezado == False:
            if celda.value != 'SI' and alumnos[f"F{celda.row}"].value == identificacion:
                alumnos[f"P{celda.row}"] = 'SI'
                celda.fill = PatternFill("solid", fgColor="92D050")
                wb.save("data/FIEC_%s_2018_2020.xlsx" % carrera)
                return True
        else:
            encabezado = False


def get_respondido(carrera, anio):
    wb = openpyxl.load_workbook('data/FIEC_%s_2018_2020.xlsx' %carrera)
    alumnos_computacion = wb['Hoja1']

    total = 0
    encabezado = True
    for row in alumnos_computacion.iter_rows():
        if encabezado == False:
            if row[15].value == 'SI':
                celda = row[10]
                if int(celda.value) == anio or int(celda.value) == anio - 1:
                    total += 1
        else:
            encabezado = False
    return total

def colorear_computacion(identificacion):
    wb = openpyxl.load_workbook('data/FIEC_Computacion_2018_2020.xlsx')
    alumnos_computacion = wb.get_sheet_by_name('Hoja1')

    for row in alumnos_computacion.iter_rows():
        if row[7].value == identificacion:
            for celda in enumerate(row):
                celda.fill = PatternFill("solid", fgColor="92D050")

def colorear(identificacion, carrera):
    wb = openpyxl.load_workbook('data/FIEC_%s_2018_2020.xlsx' %carrera)
    alumnos = wb.get_sheet_by_name('Hoja1')

    for row in alumnos.iter_rows():
        if row[5].value == identificacion :
            for celda in enumerate(row):
                celda.fill = PatternFill("solid", fgColor="92D050")

def get_total_encuestados(carrera, anio):
    wb = openpyxl.load_workbook('data/FIEC_%s_2018_2020.xlsx' % carrera)
    alumnos = wb['Hoja1']
    encabezado = True
    total = 0
    for row in alumnos.iter_rows():
        if encabezado == False:
            if int(row[10].value) == anio or int(row[10].value) == anio - 1:
                total += 1
        else:
            encabezado = False
    return total

def generar_resumen(fecha):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet['A2'] = 'Electricidad'
    sheet['A3'] = 'Electrónica y Automatización'
    sheet['A4'] = 'Computación'
    sheet['A5'] = 'Telecomunicaciones'
    sheet['A6'] = 'Telemática'
    sheet['A7'] = 'Totales'

    sheet['B1'] = 'TOTAL ENCUESTADOS 2018'
    sheet['B2'] = get_total_encuestados('Electricidad', 2018)
    sheet['B3'] = get_total_encuestados('Electronica', 2018)
    sheet['B4'] = get_total_encuestados('Computacion', 2018)
    sheet['B5'] = get_total_encuestados('Telecomunicaciones', 2018)
    sheet['B6'] = get_total_encuestados('Telematica', 2018)
    sheet['B7'] = int(sheet['B2'].value) + int(sheet['B3'].value) + int(sheet['B4'].value)\
                  + int(sheet['B5'].value) + int(sheet['B6'].value)

    sheet['C1'] = 'RESPUESTAS'
    sheet['C2'] = get_respondido('Electricidad', 2018)
    sheet['C3'] = get_respondido('Electronica', 2018)
    sheet['C4'] = get_respondido('Computacion', 2018)
    sheet['C5'] = get_respondido('Telecomunicaciones', 2018)
    sheet['C6'] = get_respondido('Telematica', 2018)
    sheet['C7'] = int(sheet['C2'].value) + int(sheet['C3'].value) + int(sheet['C4'].value)\
                  + int(sheet['C5'].value) + int(sheet['C6'].value)

    sheet['D1'] = 'PORCENTAJE'
    sheet['D2'] = (int(sheet['C2'].value) * 100) / int(sheet['B2'].value)
    sheet['D3'] = (int(sheet['C3'].value) * 100) / int(sheet['B3'].value)
    sheet['D4'] = (int(sheet['C4'].value) * 100) / int(sheet['B4'].value)
    sheet['D5'] = (int(sheet['C5'].value) * 100) / int(sheet['B5'].value)
    sheet['D6'] = (int(sheet['C6'].value) * 100) / int(sheet['B6'].value)
    sheet['D7'] = int(sheet['D2'].value) + int(sheet['D3'].value) + int(sheet['D4'].value) + int(sheet['D5'].value)\
                  + int(sheet['D6'].value)

    sheet['A13'] = 'Electricidad'
    sheet['A14'] = 'Electrónica y Automatización'
    sheet['A15'] = 'Computación'
    sheet['A16'] = 'Telecomunicaciones'
    sheet['A17'] = 'Telemática'
    sheet['A18'] = 'Totales'

    sheet['B12'] = 'TOTAL ENCUESTADOS 2020'
    sheet['B13'] = get_total_encuestados('Electricidad', 2020)
    sheet['B14'] = get_total_encuestados('Electronica', 2020)
    sheet['B15'] = get_total_encuestados('Computacion', 2020)
    sheet['B16'] = get_total_encuestados('Telecomunicaciones', 2020)
    sheet['B17'] = get_total_encuestados('Telematica', 2020)
    sheet['B18'] = int(sheet['B13'].value) + int(sheet['B14'].value) + int(sheet['B15'].value)\
                  + int(sheet['B16'].value) + int(sheet['B17'].value)

    sheet['C12'] = 'RESPUESTAS'
    sheet['C13'] = get_respondido('Electricidad', 2020)
    sheet['C14'] = get_respondido('Electronica', 2020)
    sheet['C15'] = get_respondido('Computacion', 2020)
    sheet['C16'] = get_respondido('Telecomunicaciones', 2020)
    sheet['C17'] = get_respondido('Telematica', 2020)
    sheet['C18'] = int(sheet['C13'].value) + int(sheet['C14'].value) + int(sheet['C15'].value)\
                  + int(sheet['C16'].value) + int(sheet['C17'].value)

    sheet['D12'] = 'PORCENTAJE'
    sheet['D13'] = (int(sheet['C13'].value) * 100) / int(sheet['B13'].value)
    sheet['D14'] = (int(sheet['C14'].value) * 100) / int(sheet['B14'].value)
    sheet['D15'] = (int(sheet['C15'].value) * 100) / int(sheet['B15'].value)
    sheet['D16'] = (int(sheet['C16'].value) * 100) / int(sheet['B16'].value)
    sheet['D17'] = (int(sheet['C17'].value) * 100) / int(sheet['B17'].value)
    sheet['D18'] = int(sheet['D13'].value) + int(sheet['D14'].value) + int(sheet['D15'].value)\
                  + int(sheet['D16'].value) + int(sheet['D17'].value)

    values_2018 = Reference(sheet, min_col=2, min_row=1,
                       max_col=3, max_row=6)

    values_2020 = Reference(sheet, min_col=2, min_row=12,
                            max_col=3, max_row=17)

    titles18 = Reference(sheet, min_col=1, min_row=2, max_row=6)
    titles20 = Reference(sheet, min_col=1, min_row=13, max_row=17)
    chart18 = BarChart()
    chart20 = BarChart()

    chart18.add_data(values_2018, titles_from_data=True)
    chart18.set_categories(titles18)
    chart20.add_data(values_2020, titles_from_data=True)
    chart20.set_categories(titles20)

    chart18.type = "bar"
    chart20.type = "bar"

    chart18.title = "2018"
    chart20.title = "2020"

    sheet.add_chart(chart18, "E2")
    sheet.add_chart(chart20, "N2")

    wb.save("Resumen_Reporte al %s FIEC.xlsx" %fecha)


archivo = 'Reporte 18-10-2022 FIEC.xlsx'
wb = openpyxl.load_workbook('data/%s'%archivo)
reporte = wb.get_sheet_by_name('Sheet1')
identificaciones = reporte['A']

for i, celda in enumerate(identificaciones):
        existe(celda)

fecha = archivo.split(' ')[1]
generar_resumen(fecha)
#nuevo_reporte.save("Nuevo reporte %s FIEC.xlsx")
# set_respondido('0930699541','Computacion')
# print(get_respondido('Computacion', 2018))